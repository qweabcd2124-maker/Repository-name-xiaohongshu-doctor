"""
Step 1: 数据导入与清洗
递归扫描 data/原始数据/ 所有子目录，自动识别笔记/评论文件，统一格式后写入 research.db，并导出 CSV。

Usage:
    python scripts/research/01_import_data.py
"""
from __future__ import annotations

import csv
import sqlite3
import json
import re
import sys
from pathlib import Path
from datetime import datetime
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from config import RAW_DATA_DIR, RESEARCH_DB, OUTPUT_DIR, FILE_CATEGORY_MAP, ALL_CATEGORIES


# ── Schema ──────────────────────────────────────────────────────

def create_tables(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS research_notes (
            note_id TEXT PRIMARY KEY,
            category TEXT NOT NULL,
            note_type TEXT,
            title TEXT,
            content TEXT,
            tags TEXT,
            tag_count INTEGER DEFAULT 0,
            likes INTEGER DEFAULT 0,
            collects INTEGER DEFAULT 0,
            comments_count INTEGER DEFAULT 0,
            shares INTEGER DEFAULT 0,
            engagement INTEGER DEFAULT 0,
            publish_time TEXT,
            publish_hour INTEGER,
            publish_weekday INTEGER,
            author_name TEXT,
            author_fans INTEGER DEFAULT 0,
            author_total_likes INTEGER DEFAULT 0,
            author_tier TEXT,
            image_count INTEGER DEFAULT 0,
            cover_url TEXT,
            video_duration TEXT,
            ip_location TEXT,
            source_keyword TEXT,
            title_length INTEGER DEFAULT 0,
            content_length INTEGER DEFAULT 0,
            has_emoji INTEGER DEFAULT 0,
            has_numbers INTEGER DEFAULT 0,
            title_hook_count INTEGER DEFAULT 0,
            is_viral INTEGER DEFAULT 0,
            cover_analysis TEXT,
            content_analysis TEXT,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS research_comments (
            comment_id TEXT PRIMARY KEY,
            note_id TEXT,
            user_id TEXT,
            user_name TEXT,
            content TEXT,
            likes INTEGER DEFAULT 0,
            comment_time TEXT,
            ip_location TEXT,
            sub_comment_count INTEGER DEFAULT 0,
            parent_comment_id TEXT,
            parent_comment_content TEXT,
            parent_user_name TEXT,
            reply_to_comment_id TEXT,
            reply_to_content TEXT,
            reply_to_user_name TEXT,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_rn_category ON research_notes(category)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_rn_viral ON research_notes(is_viral)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_rn_engagement ON research_notes(engagement)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_rc_note ON research_comments(note_id)")


# ── Helpers ─────────────────────────────────────────────────────

def safe_int(val) -> int:
    if val is None or val == "":
        return 0
    try:
        return int(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def detect_category(filename: str) -> str | None:
    name = Path(filename).stem
    for prefix, cat in FILE_CATEGORY_MAP.items():
        if prefix in name:
            return cat
    return None


def detect_category_from_keyword(keyword: str) -> str | None:
    """从第二批数据的'来源关键词'列推断品类"""
    if not keyword:
        return None
    kw = str(keyword).strip()
    mapping = {
        "美食": "food", "做饭": "food", "食谱": "food", "烘焙": "food",
        "穿搭": "fashion", "时尚": "fashion", "服装": "fashion", "outfit": "fashion", "显瘦": "fashion",
        "科技": "tech", "手机": "tech", "数码": "tech", "电脑": "tech", "AI": "tech", "测评": "tech",
        "旅游": "travel", "旅行": "travel", "景点": "travel", "小众旅行": "travel",
        "美妆": "beauty", "护肤": "beauty", "化妆": "beauty",
        "健身": "fitness", "运动": "fitness", "减肥": "fitness",
        "生活": "lifestyle", "日常": "lifestyle",
        "家居": "home", "装修": "home", "家装": "home",
    }
    for k, v in mapping.items():
        if k in kw:
            return v
    return None


def detect_category_from_content(title: str, tags: str, content: str) -> str | None:
    """从标题/话题/内容推断品类（兜底）"""
    text = (title + " " + tags + " " + content[:200]).lower()
    rules = [
        ("food", ["美食", "食谱", "做饭", "烹饪", "烘焙", "食材", "好吃"]),
        ("fashion", ["穿搭", "outfit", "时尚", "服装", "搭配", "显瘦"]),
        ("tech", ["科技", "数码", "手机", "电脑", "AI", "测评", "编程", "推荐系统"]),
        ("travel", ["旅游", "旅行", "景点", "攻略", "打卡"]),
        ("beauty", ["美妆", "护肤", "化妆", "口红", "防晒"]),
        ("fitness", ["健身", "运动", "减肥", "瑜伽", "跑步"]),
        ("lifestyle", ["生活", "日常", "vlog", "觉醒", "追星", "闺蜜"]),
        ("home", ["家居", "装修", "家装", "收纳"]),
    ]
    for cat, keywords in rules:
        if any(k in text for k in keywords):
            return cat
    return "lifestyle"  # 最终兜底


def is_comment_file(headers: list[str]) -> bool:
    return "评论ID" in headers and "评论内容" in headers


def parse_tags(raw_topics) -> list[str]:
    if not raw_topics:
        return []
    cleaned = re.sub(r'\[话题\]', '', str(raw_topics))
    tags = re.split(r'[、，,\n]', cleaned)
    return [t.strip().strip('#') for t in tags if t.strip().strip('#')]


def detect_emoji(text: str) -> bool:
    emoji_pattern = re.compile(
        "[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF"
        "\U0001F900-\U0001F9FF\U00002702-\U000027B0✨🔥💚🧋🍱‼️⭐📸📊👍👎]",
        flags=re.UNICODE
    )
    return bool(emoji_pattern.search(text or ""))


def count_hooks(title: str) -> int:
    hooks = 0
    if re.search(r'\d+', title):
        hooks += 1
    if re.search(r'[！!？?]', title):
        hooks += 1
    if re.search(r'[｜|]', title):
        hooks += 1
    if re.search(r'[✨🔥‼️⭐💯]', title):
        hooks += 1
    if re.search(r'(必|绝了|太|超|巨|神仙|宝藏|救命)', title):
        hooks += 1
    return hooks


def classify_author_tier(total_likes: int) -> str:
    if total_likes < 5000:
        return "nano"
    elif total_likes < 50000:
        return "micro"
    elif total_likes < 500000:
        return "mid"
    else:
        return "macro"


def parse_note_type(raw) -> str:
    if not raw:
        return "image"
    r = str(raw).strip()
    if "视频" in r or "video" in r.lower():
        return "video"
    return "image"


def parse_datetime(val):
    if val is None:
        return None, None, None
    if isinstance(val, datetime):
        return val.isoformat(), val.hour, val.weekday()
    try:
        dt = datetime.fromisoformat(str(val).strip())
        return dt.isoformat(), dt.hour, dt.weekday()
    except (ValueError, TypeError):
        return str(val), None, None


def get_field(data: dict, *keys, default=None):
    """从 dict 中按多个可能的 key 取值"""
    for k in keys:
        if k in data and data[k] is not None:
            return data[k]
    return default


# ── Note processing ─────────────────────────────────────────────

def process_note_xlsx(filepath: str, default_category: str | None) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []

    for row in rows[1:]:
        data = dict(zip(headers, row))

        note_id = str(get_field(data, "笔记ID") or "").strip()
        if not note_id:
            continue

        title = str(get_field(data, "笔记标题") or "").strip()
        content = str(get_field(data, "笔记内容") or "").strip()

        # 品类：优先来源关键词 → 文件名推断 → 内容推断
        category = (
            detect_category_from_keyword(get_field(data, "来源关键词"))
            or default_category
            or detect_category_from_content(title, str(get_field(data, "笔记话题") or ""), content)
        )
        tags = parse_tags(get_field(data, "笔记话题"))
        likes = safe_int(get_field(data, "点赞量"))
        collects = safe_int(get_field(data, "收藏量"))
        comments_count = safe_int(get_field(data, "评论量"))
        shares = safe_int(get_field(data, "分享量"))
        engagement = likes + collects + comments_count + shares
        author_fans = safe_int(get_field(data, "粉丝数"))
        author_total = safe_int(get_field(data, "获赞与收藏"))
        image_count = safe_int(get_field(data, "图片数量"))

        # 发布时间：有些文件分为"发布日期"和"发布时间"两列
        pub_raw = get_field(data, "发布时间", "发布日期")
        publish_time, publish_hour, publish_weekday = parse_datetime(pub_raw)

        records.append({
            "note_id": note_id,
            "category": category,
            "note_type": parse_note_type(get_field(data, "笔记类型")),
            "title": title,
            "content": content,
            "tags": json.dumps(tags, ensure_ascii=False),
            "tag_count": len(tags),
            "likes": likes,
            "collects": collects,
            "comments_count": comments_count,
            "shares": shares,
            "engagement": engagement,
            "publish_time": publish_time,
            "publish_hour": publish_hour,
            "publish_weekday": publish_weekday,
            "author_name": str(get_field(data, "博主昵称") or "").strip(),
            "author_fans": author_fans,
            "author_total_likes": author_total,
            "author_tier": classify_author_tier(author_total or author_fans),
            "image_count": image_count,
            "cover_url": str(get_field(data, "笔记封面链接") or "").strip(),
            "video_duration": str(get_field(data, "笔记视频时长") or "").strip() or None,
            "ip_location": str(get_field(data, "IP地址") or "").strip() or None,
            "source_keyword": str(get_field(data, "来源关键词") or "").strip() or None,
            "title_length": len(title),
            "content_length": len(content),
            "has_emoji": 1 if detect_emoji(title + content) else 0,
            "has_numbers": 1 if re.search(r'\d+', title) else 0,
            "title_hook_count": count_hooks(title),
        })

    return records


# ── Comment processing ──────────────────────────────────────────

def process_comment_xlsx(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []

    for row in rows[1:]:
        data = dict(zip(headers, row))
        comment_id = str(get_field(data, "评论ID") or "").strip()
        if not comment_id:
            continue

        records.append({
            "comment_id": comment_id,
            "note_id": str(get_field(data, "笔记ID") or "").strip(),
            "user_id": str(get_field(data, "用户ID") or "").strip(),
            "user_name": str(get_field(data, "用户名称") or "").strip(),
            "content": str(get_field(data, "评论内容") or "").strip(),
            "likes": safe_int(get_field(data, "点赞量")),
            "comment_time": str(get_field(data, "评论时间") or "").strip() or None,
            "ip_location": str(get_field(data, "IP地址") or "").strip() or None,
            "sub_comment_count": safe_int(get_field(data, "子评论数")),
            "parent_comment_id": str(get_field(data, "一级评论ID") or "").strip() or None,
            "parent_comment_content": str(get_field(data, "一级评论内容") or "").strip() or None,
            "parent_user_name": str(get_field(data, "一级评论用户名称") or "").strip() or None,
            "reply_to_comment_id": str(get_field(data, "引用的评论ID") or "").strip() or None,
            "reply_to_content": str(get_field(data, "引用的评论内容") or "").strip() or None,
            "reply_to_user_name": str(get_field(data, "引用的用户名称") or "").strip() or None,
        })

    return records


# ── Viral threshold ─────────────────────────────────────────────

def compute_viral_threshold(cursor):
    for cat in ALL_CATEGORIES:
        cursor.execute(
            "SELECT engagement FROM research_notes WHERE category=? ORDER BY engagement",
            (cat,)
        )
        vals = [r[0] for r in cursor.fetchall()]
        if not vals:
            continue
        p90_idx = int(len(vals) * 0.9)
        threshold = vals[min(p90_idx, len(vals) - 1)]
        cursor.execute(
            "UPDATE research_notes SET is_viral=1 WHERE category=? AND engagement>=?",
            (cat, threshold)
        )
        viral = len(vals) - p90_idx
        print(f"  [{cat}] {len(vals)} 条, P90={threshold}, 爆款 {viral} 条")


# ── CSV export ──────────────────────────────────────────────────

def export_csv(conn):
    csv_dir = OUTPUT_DIR / "csv"
    csv_dir.mkdir(exist_ok=True)

    # Export notes
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM research_notes")
    cols = [d[0] for d in cursor.description]
    rows = cursor.fetchall()
    notes_csv = csv_dir / "research_notes.csv"
    with open(notes_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(cols)
        w.writerows(rows)
    print(f"  → {notes_csv} ({len(rows)} rows)")

    # Export per-category
    for cat in ALL_CATEGORIES:
        cursor.execute("SELECT * FROM research_notes WHERE category=?", (cat,))
        cat_rows = cursor.fetchall()
        if cat_rows:
            cat_csv = csv_dir / f"notes_{cat}.csv"
            with open(cat_csv, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(cols)
                w.writerows(cat_rows)

    # Export comments
    cursor.execute("SELECT * FROM research_comments")
    cols_c = [d[0] for d in cursor.description]
    rows_c = cursor.fetchall()
    if rows_c:
        comments_csv = csv_dir / "research_comments.csv"
        with open(comments_csv, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(cols_c)
            w.writerows(rows_c)
        print(f"  → {comments_csv} ({len(rows_c)} rows)")


# ── Main ────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Step 1: 数据导入与清洗")
    print("=" * 60)

    # Ensure DB directory exists
    RESEARCH_DB.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(RESEARCH_DB)
    cursor = conn.cursor()

    # Drop old tables for clean import
    cursor.execute("DROP TABLE IF EXISTS research_notes")
    cursor.execute("DROP TABLE IF EXISTS research_comments")
    create_tables(cursor)

    total_notes = 0
    total_comments = 0
    seen_comment_ids = set()

    # Recursively find all xlsx files
    xlsx_files = sorted(RAW_DATA_DIR.rglob("*.xlsx"))
    print(f"\n找到 {len(xlsx_files)} 个 xlsx 文件\n")

    for f in xlsx_files:
        if f.name.startswith("~$"):  # skip temp files
            continue

        # Read headers to determine file type
        try:
            wb = openpyxl.load_workbook(str(f), read_only=True)
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()
        except Exception as e:
            print(f"  [跳过] {f.name} — 无法读取: {e}")
            continue

        if not first_row:
            continue
        headers = [str(h).strip() if h else "" for h in first_row]

        if is_comment_file(headers):
            # ── Comment file ──
            records = process_comment_xlsx(str(f))
            imported = 0
            for r in records:
                if r["comment_id"] in seen_comment_ids:
                    continue
                seen_comment_ids.add(r["comment_id"])
                try:
                    cursor.execute("""
                        INSERT OR IGNORE INTO research_comments (
                            comment_id, note_id, user_id, user_name, content, likes,
                            comment_time, ip_location, sub_comment_count,
                            parent_comment_id, parent_comment_content, parent_user_name,
                            reply_to_comment_id, reply_to_content, reply_to_user_name
                        ) VALUES (
                            :comment_id, :note_id, :user_id, :user_name, :content, :likes,
                            :comment_time, :ip_location, :sub_comment_count,
                            :parent_comment_id, :parent_comment_content, :parent_user_name,
                            :reply_to_comment_id, :reply_to_content, :reply_to_user_name
                        )
                    """, r)
                    imported += 1
                except Exception as e:
                    print(f"    评论错误: {r['comment_id']} — {e}")
            total_comments += imported
            print(f"  📝 评论 {f.name}: {imported} 条")
        else:
            # ── Note file ──
            category = detect_category(f.name)
            records = process_note_xlsx(str(f), category)
            imported = 0
            for r in records:
                try:
                    cursor.execute("""
                        INSERT OR REPLACE INTO research_notes (
                            note_id, category, note_type, title, content, tags, tag_count,
                            likes, collects, comments_count, shares, engagement,
                            publish_time, publish_hour, publish_weekday,
                            author_name, author_fans, author_total_likes, author_tier,
                            image_count, cover_url, video_duration, ip_location, source_keyword,
                            title_length, content_length, has_emoji, has_numbers, title_hook_count
                        ) VALUES (
                            :note_id, :category, :note_type, :title, :content, :tags, :tag_count,
                            :likes, :collects, :comments_count, :shares, :engagement,
                            :publish_time, :publish_hour, :publish_weekday,
                            :author_name, :author_fans, :author_total_likes, :author_tier,
                            :image_count, :cover_url, :video_duration, :ip_location, :source_keyword,
                            :title_length, :content_length, :has_emoji, :has_numbers, :title_hook_count
                        )
                    """, r)
                    imported += 1
                except Exception as e:
                    print(f"    笔记错误: {r.get('note_id')} — {e}")

            cat_label = category or "未分类"
            total_notes += imported
            if imported > 0:
                print(f"  📄 笔记 {f.name} → [{cat_label}]: {imported} 条")
            elif records:
                print(f"  [跳过] {f.name} — 无法确定品类, {len(records)} 条未导入")
            else:
                print(f"  [跳过] {f.name} — 无有效数据")

    # 计算爆款阈值
    print("\n计算爆款阈值 (P90)...")
    compute_viral_threshold(cursor)

    conn.commit()

    # 统计报告
    print(f"\n{'='*60}")
    print("导入统计")
    print(f"{'='*60}")
    cursor.execute("SELECT category, COUNT(*), SUM(is_viral) FROM research_notes GROUP BY category")
    for cat, total, viral in cursor.fetchall():
        print(f"  {cat:12s}: {total:4d} 条 ({viral or 0} 爆款)")
    cursor.execute("SELECT COUNT(*) FROM research_notes")
    n = cursor.fetchone()[0]
    print(f"\n  笔记总计: {n} 条")
    print(f"  评论总计: {total_comments} 条")

    # 导出 CSV
    print(f"\n导出 CSV...")
    export_csv(conn)

    conn.close()
    print(f"\n数据库: {RESEARCH_DB}")


if __name__ == "__main__":
    main()
